from collections import deque
from KNN import KNN
from openpyxl import load_workbook
import numpy as np
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt

# training data
wb = load_workbook(filename='Data.xlsx')
ws = wb.active
colA = ws['A']
colB = ws['B']
X_set = np.array([])
a = deque()
for X in colA:
    T = X.value.split(' ')
    if len(T) > 2:
        T.pop(2)
    R = map(int, T)
    E = list(R)
    a.append(E)
X_set = np.array(a)
Y_set = np.array([])
for Y in colB:
    Y_set = np.append(Y_set, Y.value)

#Test data
wb_test = load_workbook(filename='Input.xlsx')
ws_test = wb_test.active
colA_test = ws_test['A']
X_test = np.array([])
a.clear()
for X in colA_test:
    T = X.value.split(' ')
    if len(T) > 2:
        T.pop(2)
    R = map(int, T)
    E = list(R)
    a.append(E)
X_test = np.array(a)
# Number of clusters
ClusterNum = 2
# Clustering
kmeans = KMeans(n_clusters=ClusterNum, random_state=0)
cluster_labels = kmeans.fit_predict(X_set)

# Getting the number of iterations
iterations = kmeans.n_iter_
print("Number of iterations:", iterations)

# Optionally, plot the cluster centroids
cluster_centers = kmeans.cluster_centers_

# Figure 1
# Create a scatter plot for each cluster
plt.figure('KNN++', figsize=[3,3])
for cluster_num in range(ClusterNum):
    plt.scatter(X_set[cluster_labels == cluster_num, 0], X_set[cluster_labels == cluster_num, 1], label=f'Cluster {cluster_num + 1}')
plt.scatter(cluster_centers[:, 0], cluster_centers[:, 1], s=100, c='black', marker='X', label='Centroids')
plt.xlabel('X')
plt.ylabel('Y')
plt.legend()
plt.title('Cluster Visualization (' + str(iterations) + ' iterations)')

#Figure 2
plt.figure('KNNRand')
# Figure 3
plt.figure('DBSCAN')
plt.show()

# Create and fit the improved KNN model with cluster labels
knn = KNN(k=1)
knn.fit(X_set, cluster_labels)

# Predict cluster labels for the test data
cluster_predictions = knn.predict(X_test)

# Output cluster predictions
print("Cluster Predictions:")
for i, cluster_label in enumerate(cluster_predictions):
    cluster_visual = f"[{cluster_label}]"
    print(f"Data Point {i + 1}: {cluster_visual}", end=' ')
    if (i + 1) % 10 == 0:
        print()  # Start a new line after every 10 data points

# Print a legend
print("\nCluster Legend:")
for cluster_label in set(cluster_predictions):
    cluster_visual = f"[{cluster_label}]"
    print(f"Cluster {cluster_label}: {cluster_visual}")

